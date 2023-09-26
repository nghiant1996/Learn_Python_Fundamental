import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}

print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inv = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

    # add value for total inventory price
    product_list.cell(product_row, 5).value = inv * price

    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += inv
    else:
        print("adding a new supplier")
        products_per_supplier[supplier_name] = inv

inv_file.save("new_inventory.xlxs")
print(products_per_supplier)